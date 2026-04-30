#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_dashboard.py
================
Gera o arquivo dashboard_mege.html a partir de tickets_etiquetados_mege_completo.xlsx.

Uso:
    pip install openpyxl
    python gen_dashboard.py

Saída:
    dashboard_mege.html  (~4 MB, standalone, abre direto no navegador)

─────────────────────────────────────────────────────────
PROCESSO DE INVESTIGAÇÃO E DECISÕES DE DESIGN
─────────────────────────────────────────────────────────

1. DIAGNÓSTICO INICIAL
   O arquivo fonte tinha 4.626 linhas e 22 colunas. Campos relevantes:
   - Etiquetados por IA: tipo_solicitacao, material_referenciado, carreira,
     materia, confianca, nota_ambiguidade
   - Classificação original do sistema: tipo_suporte (7 categorias humanas,
     diferente da taxonomia da IA)
   - Dados do aluno: nome_aluno, email_aluno (não embarcados no dashboard
     por privacidade — apenas id_aluno é preservado via campo `id`)

2. IDENTIFICAÇÃO DE DUPLICATAS
   Ao cruzar os IDs (UUID v4), descobriu-se que o banco continha 408 linhas
   duplicadas — o mesmo ticket aparecia 2x, 3x ou até 4x no arquivo.

   Diagnóstico detalhado:
     Total de linhas brutas : 4.626
     IDs únicos             : 4.218
     IDs duplicados         :   398  (aparecem 2x ou mais)
     Linhas redundantes     :   408  (removidas)

   Causa provável: exportação com JOIN sem DISTINCT ou múltiplas inserções
   no banco de origem.

   Solução: `seen_ids = set()` — na primeira ocorrência de cada ID o registro
   é processado; nas seguintes, `continue` pula para a próxima linha.
   Isso garante que o dashboard sempre reflita 4.218 tickets únicos.

   Para auditar os duplicados isoladamente, use o script abaixo:

       from collections import Counter
       import openpyxl
       wb = openpyxl.load_workbook('tickets_etiquetados_mege_completo.xlsx', read_only=True)
       ws = wb.active
       headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
       ids = [dict(zip(headers, r)).get('id','') for r in ws.iter_rows(min_row=2, values_only=True)]
       duplicados = {k: v for k, v in Counter(ids).items() if v > 1}
       # duplicados = {uuid: n_ocorrencias, ...}  →  398 entradas

3. CAMPOS DERIVADOS
   a) `fase` — não existia no banco. Derivada do campo `turma` por matching
      de substrings (case-insensitive). Ver função derive_fase() abaixo.

   b) `turma_clean` — nome original tinha conteúdo entre parênteses muito
      longo (ex: "TJSP 192, 2ª Fase (Turma 1: Materiais, videoaulas e
      correções por magistrados)"). Removido via regex e truncado em 50 chars.

   c) `dia`, `mes`, `semana`, `dow` — extraídos de data_criacao para
      alimentar os gráficos de evolução temporal e dia da semana.

4. CAMPO tipo_suporte
   Descoberto durante investigação: além das tags de IA (tipo_solicitacao),
   o banco tinha uma classificação humana original (tipo_suporte) com
   categorias diferentes. Adicionado como filtro independente em todas as
   abas para permitir cruzamento entre classificação humana e IA.

5. ESTRUTURA DOS ARRAYS EMBARCADOS
   Para manter o HTML abaixo de 6 MB, os dados são divididos em dois arrays:

   FD (Filter Data) — 12 campos por linha, ~700 KB:
     [0] data          [1] tipo_solicitacao  [2] carreira
     [3] fase          [4] materia           [5] material_referenciado
     [6] turma         [7] status            [8] mes
     [9] semana        [10] dow              [11] tipo_suporte

   DD (Detail Data) — 17 campos por linha, ~3.3 MB:
     [0] data          [1] titulo            [2] tipo_solicitacao
     [3] carreira      [4] fase              [5] materia
     [6] material      [7] turma             [8] status
     [9] coordenador   [10] confianca        [11] descricao (max 800 chars)
     [12] resposta     [13] nota_ambiguidade [14] mes
     [15] tipo_suporte [16] id
"""

import openpyxl, json, re
from datetime import datetime

wb = openpyxl.load_workbook('tickets_etiquetados_mege_completo.xlsx', read_only=True)
ws = wb.active


def clean_turma(t):
    """Remove conteúdo entre parênteses do nome da turma e trunca em 50 chars.
    Ex: 'TJSP 192, 2ª Fase (Turma 1: Materiais...)' → 'TJSP 192, 2ª Fase'
    """
    if not t: return 'Sem turma'
    t = re.sub(r'\s*\([^)]*\)', '', str(t)).strip()
    return t[:50] if len(t) > 50 else t


def derive_fase(turma):
    """Deriva a fase do curso a partir do nome da turma.
    O campo 'fase' não existe no banco — é inferido por matching de substrings.
    Cobre variações de encoding (ª/°/a) e com/sem hífen.
    """
    if not turma: return 'Geral'
    t = str(turma).lower()
    if any(x in t for x in ['2ª fase','2a fase','2° fase',' 2 fase','segunda fase',
                              '2\xaa fase','2\xb0 fase','2ª fase','2° fase']):
        return '2ª Fase'
    if any(x in t for x in ['pré-edital','pre-edital','pré edital','pre edital',
                              'pr\xe9-edital','pr\xe9 edital']): return 'Pré-Edital'
    if 'clube' in t: return 'Clube'
    if any(x in t for x in ['pós-edital','pos-edital','pós edital','pos edital',
                              'p\xf3s-edital','p\xf3s edital']): return 'Pós-Edital'
    if any(x in t for x in ['até passar','ate passar',
                              'at\xe9 passar']): return 'Até Passar'
    if any(x in t for x in ['1ª fase','1a fase','primeira fase',
                              '1\xaa fase']): return '1ª Fase'
    if 'reta final' in t: return 'Reta Final'
    return 'Geral'


headers = None
rows_filter = []   # array FD — dados para filtros e gráficos
rows_full   = []   # array DD — dados completos para o Explorador

# ─── DEDUPLICAÇÃO ────────────────────────────────────────────────────────────
# O banco continha 408 linhas duplicadas (mesmo UUID aparecendo 2-4x).
# seen_ids rastreia IDs já processados; registros repetidos são ignorados.
# Resultado: 4.626 linhas brutas → 4.218 tickets únicos.
seen_ids = set()

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
        continue

    r = dict(zip(headers, row))

    # ── Deduplicação por ID ──────────────────────────────────────────────────
    row_id = str(r.get('id','') or '')
    if row_id and row_id in seen_ids:
        continue          # duplicata — pula
    if row_id:
        seen_ids.add(row_id)

    # ── Limpeza e derivação de turma/fase ───────────────────────────────────
    turma_raw   = r.get('turma','') or ''
    turma_clean = clean_turma(turma_raw)
    fase        = derive_fase(turma_raw)

    # ── Derivação de campos de data ──────────────────────────────────────────
    dc = r.get('data_criacao','')
    if dc:
        try:
            dt     = dc if isinstance(dc, datetime) else datetime.fromisoformat(str(dc)[:19])
            data   = dt.strftime('%Y-%m-%d')
            mes    = dt.strftime('%Y-%m')
            semana = dt.strftime('%Y-W%W')
            dow    = dt.weekday()          # 0=Segunda … 6=Domingo
        except:
            data = mes = semana = ''; dow = -1
    else:
        data = mes = semana = ''; dow = -1

    status       = str(r.get('status','') or '').strip()
    tipo_sup_raw = str(r.get('tipo_suporte','') or '').strip().rstrip()

    # ── Array FD (filtros e gráficos) ────────────────────────────────────────
    rows_filter.append([
        data,
        str(r.get('tipo_solicitacao','') or ''),  # [1]
        str(r.get('carreira','') or ''),           # [2]
        fase,                                      # [3]
        str(r.get('materia','') or ''),            # [4]
        str(r.get('material_referenciado','') or ''),  # [5]
        turma_clean,                               # [6]
        status,                                    # [7]
        mes,                                       # [8]
        semana,                                    # [9]
        dow,                                       # [10]
        tipo_sup_raw                               # [11]
    ])

    # ── Array DD (explorador — dados completos) ──────────────────────────────
    titulo = str(r.get('titulo','') or '')[:200]
    # Descrições e respostas chegam com HTML — strip de tags + truncagem em 800 chars
    desc   = re.sub(r'<[^>]+>',' ', str(r.get('descricao','') or '')).strip()[:800]
    resp   = re.sub(r'<[^>]+>',' ', str(r.get('reposta','')   or '')).strip()[:800]
    coord  = str(r.get('coodenador_responsavel','') or '')
    conf   = r.get('confianca', None)
    try:   conf = round(float(conf), 2) if conf is not None else None
    except: conf = None
    nota   = str(r.get('nota_ambiguidade','') or '')[:300] if r.get('nota_ambiguidade') else ''
    rows_full.append([
        data,
        titulo,
        str(r.get('tipo_solicitacao','') or ''),
        str(r.get('carreira','') or ''),
        fase,
        str(r.get('materia','') or ''),
        str(r.get('material_referenciado','') or ''),
        turma_clean,
        status,
        coord,
        conf,
        desc,
        resp,
        nota,
        mes,
        tipo_sup_raw,                        # [15]
        str(r.get('id','') or '')            # [16]
    ])

FILTER_JSON = json.dumps(rows_filter, ensure_ascii=False, separators=(',',':'))
FULL_JSON = json.dumps(rows_full, ensure_ascii=False, separators=(',',':'))

HTML = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Dashboard Suporte MEGE</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Syne:wght@600;700;800&display=swap" rel="stylesheet"/>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#f8f9fc;--card:#fff;--border:#e2e8f0;--text:#1a202c;--muted:#718096;
  --blue:#09408C;--orange:#F58220;--green:#029A42;--purple:#6400FC;
  --red:#E53E3E;--teal:#0694A2;--pink:#D53F8C;--yellow:#D69E2E;
  --shadow:0 1px 3px rgba(0,0,0,.08),0 1px 2px rgba(0,0,0,.05);
  --shadow-md:0 4px 6px rgba(0,0,0,.07),0 2px 4px rgba(0,0,0,.05);
}
body{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.5}
h1,h2,h3,.syne{font-family:'Syne',sans-serif}

/* NAV */
.topbar{background:var(--blue);padding:0 24px;display:flex;align-items:center;gap:24px;box-shadow:0 2px 8px rgba(9,64,140,.3)}
.topbar-logo{color:#fff;font-family:'Syne',sans-serif;font-size:20px;font-weight:800;letter-spacing:.05em;padding:14px 0;white-space:nowrap}
.topbar-logo span{color:var(--orange)}
.tabs{display:flex;gap:0;margin-left:auto}
.tab-btn{background:none;border:none;color:rgba(255,255,255,.7);padding:16px 20px;cursor:pointer;font-family:'Inter',sans-serif;font-size:13px;font-weight:500;border-bottom:3px solid transparent;transition:all .2s;white-space:nowrap}
.tab-btn:hover{color:#fff;background:rgba(255,255,255,.08)}
.tab-btn.active{color:#fff;border-bottom-color:var(--orange)}

/* LAYOUT */
.content{display:none;padding:24px;max-width:1400px;margin:0 auto}
.content.active{display:block}

/* FILTER BAR */
.filter-bar{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px 20px;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end;margin-bottom:20px;box-shadow:var(--shadow)}
.filter-group{display:flex;flex-direction:column;gap:4px}
.filter-group label{font-size:11px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em}
.filter-group select,.filter-group input{border:1px solid var(--border);border-radius:8px;padding:7px 10px;font-size:13px;background:#fff;color:var(--text);min-width:140px;outline:none;font-family:'Inter',sans-serif}
.filter-group select:focus,.filter-group input:focus{border-color:var(--blue);box-shadow:0 0 0 3px rgba(9,64,140,.1)}
.btn-clear{background:none;border:1px solid var(--border);border-radius:8px;padding:7px 14px;cursor:pointer;font-size:13px;color:var(--muted);transition:all .2s;align-self:flex-end}
.btn-clear:hover{border-color:var(--red);color:var(--red)}

/* KPIs */
.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:20px}
.kpi{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px;box-shadow:var(--shadow);border-left:4px solid var(--blue)}
.kpi.orange{border-left-color:var(--orange)}
.kpi.green{border-left-color:var(--green)}
.kpi.purple{border-left-color:var(--purple)}
.kpi-label{font-size:11px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px}
.kpi-value{font-family:'Syne',sans-serif;font-size:32px;font-weight:700;color:var(--text);line-height:1}
.kpi-sub{font-size:12px;color:var(--muted);margin-top:4px}

/* GRID LAYOUTS */
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}
.grid-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px}
.grid-4{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:16px}
.col-span-2{grid-column:span 2}

/* CARDS */
.card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px;box-shadow:var(--shadow)}
.card-title{font-family:'Syne',sans-serif;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--text);margin-bottom:16px;display:flex;align-items:center;gap:8px}
.card-title .badge-count{background:var(--bg);border:1px solid var(--border);border-radius:20px;padding:2px 8px;font-size:11px;font-weight:600;color:var(--muted)}

/* HORIZONTAL BAR CHART */
.hbar-list{display:flex;flex-direction:column;gap:8px}
.hbar-item{display:grid;grid-template-columns:170px 1fr 60px 48px;align-items:center;gap:8px;cursor:default}
.hbar-item.clickable{cursor:pointer}
.hbar-item.clickable:hover .hbar-label{color:var(--blue)}
.hbar-label{font-size:12px;font-weight:500;color:var(--text);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.hbar-track{background:#f1f5f9;border-radius:4px;height:10px;overflow:hidden}
.hbar-fill{height:100%;border-radius:4px;transition:width .4s ease}
.hbar-val{font-size:12px;font-weight:600;color:var(--text);text-align:right}
.hbar-pct{font-size:11px;color:var(--muted);text-align:right}

/* PARETO MARKER */
.pareto-line{border-top:2px dashed var(--orange);margin:6px 0 2px;position:relative}
.pareto-line::before{content:'80% do volume';position:absolute;right:0;top:-18px;font-size:10px;color:var(--orange);font-weight:600;background:#fff;padding:0 4px}

/* CHART CONTAINER */
.chart-wrap{position:relative;width:100%}
.toggle-group{display:flex;gap:0;background:#f1f5f9;border-radius:8px;padding:3px;margin-bottom:12px;align-self:flex-start}
.toggle-btn{background:none;border:none;padding:5px 12px;border-radius:6px;cursor:pointer;font-size:12px;font-weight:500;color:var(--muted);transition:all .2s}
.toggle-btn.active{background:#fff;color:var(--blue);box-shadow:0 1px 3px rgba(0,0,0,.1)}

/* FASE CARDS - Aba 2 */
.fase-cards{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:20px}
.fase-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px;box-shadow:var(--shadow)}
.fase-card-name{font-family:'Syne',sans-serif;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px}
.fase-card-vol{font-size:22px;font-weight:700;color:var(--text);margin-bottom:10px}
.fase-mini-bar{margin-bottom:6px}
.fase-mini-label{display:flex;justify-content:space-between;font-size:11px;margin-bottom:2px}
.fase-mini-track{background:#f1f5f9;border-radius:3px;height:6px;overflow:hidden}
.fase-mini-fill{height:100%;border-radius:3px}
.fase-material{font-size:11px;color:var(--muted);margin-top:8px;border-top:1px solid var(--border);padding-top:8px}

/* INSIGHT CARDS */
.insight-cards{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px}
.insight-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px;box-shadow:var(--shadow)}
.insight-icon{font-size:22px;margin-bottom:8px}
.insight-title{font-size:12px;font-weight:700;color:var(--text);margin-bottom:4px}
.insight-text{font-size:12px;color:var(--muted);line-height:1.5}

/* HEATMAP */
.heatmap-wrap{overflow-x:auto}
.heatmap{border-collapse:collapse;font-size:11px;width:100%}
.heatmap th{padding:6px 8px;font-weight:600;text-align:left;color:var(--muted);white-space:nowrap;background:var(--bg)}
.heatmap td{padding:5px 8px;text-align:center;border:1px solid var(--border);min-width:55px}
.heatmap .row-label{text-align:left;font-weight:500;white-space:nowrap;background:var(--bg)}

/* BADGES */
.badge{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:600;white-space:nowrap}

/* EXPLORER TABLE */
.tbl-wrap{overflow-x:auto}
table.tbl{width:100%;border-collapse:collapse;font-size:13px}
table.tbl th{padding:10px 12px;text-align:left;font-weight:600;color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.04em;border-bottom:2px solid var(--border);background:var(--bg);white-space:nowrap}
table.tbl td{padding:10px 12px;border-bottom:1px solid var(--border);vertical-align:middle}
table.tbl tr:hover td{background:#f8faff;cursor:pointer}
table.tbl tr.selected td{background:#eff6ff}
.tbl-titulo{max-width:280px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-weight:500}

/* SIDE PANEL */
.side-panel{position:fixed;top:0;right:-440px;width:420px;height:100vh;background:var(--card);border-left:1px solid var(--border);box-shadow:-4px 0 20px rgba(0,0,0,.1);z-index:100;transition:right .3s ease;overflow-y:auto;display:flex;flex-direction:column}
.side-panel.open{right:0}
.side-panel-header{padding:20px;border-bottom:1px solid var(--border);display:flex;align-items:flex-start;justify-content:space-between;gap:12px;position:sticky;top:0;background:var(--card);z-index:1}
.side-panel-titulo{font-weight:700;font-size:15px;line-height:1.4;flex:1}
.btn-close{background:none;border:1px solid var(--border);border-radius:8px;width:32px;height:32px;cursor:pointer;font-size:16px;color:var(--muted);display:flex;align-items:center;justify-content:center;flex-shrink:0;transition:all .2s}
.btn-close:hover{border-color:var(--red);color:var(--red)}
.side-panel-body{padding:20px;flex:1}
.sp-badges{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:16px}
.sp-meta{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:16px}
.sp-meta-item{background:var(--bg);border-radius:8px;padding:10px}
.sp-meta-label{font-size:10px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;margin-bottom:2px}
.sp-meta-value{font-size:13px;font-weight:600;color:var(--text)}
.sp-section{margin-bottom:14px}
.sp-section-label{font-size:11px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px}
.sp-box{background:var(--bg);border-radius:8px;padding:12px;font-size:13px;line-height:1.6;color:var(--text)}
.sp-box.green{background:#f0fdf4;border:1px solid #bbf7d0}
.sp-box.yellow{background:#fffbeb;border:1px solid #fde68a}

/* PAGINATION */
.pagination{display:flex;align-items:center;gap:8px;margin-top:16px;justify-content:center}
.pg-btn{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:6px 12px;cursor:pointer;font-size:13px;color:var(--text);transition:all .2s}
.pg-btn:hover:not(:disabled){border-color:var(--blue);color:var(--blue)}
.pg-btn:disabled{opacity:.4;cursor:default}
.pg-btn.active{background:var(--blue);border-color:var(--blue);color:#fff}
.pg-info{font-size:12px;color:var(--muted)}

/* RESULT COUNT */
.result-count{font-size:13px;color:var(--muted);margin-bottom:12px}
.result-count strong{color:var(--text)}

/* SECTION HEADER */
.section-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px}
.section-title{font-family:'Syne',sans-serif;font-size:18px;font-weight:700;color:var(--text)}
.section-sub{font-size:13px;color:var(--muted);margin-top:2px}

/* CARREIRA FILTER */
.carreira-pills{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:20px}
.pill{background:var(--bg);border:1px solid var(--border);border-radius:20px;padding:6px 14px;cursor:pointer;font-size:13px;font-weight:500;color:var(--muted);transition:all .2s}
.pill.active{color:#fff;border-color:transparent}
.pill:hover:not(.active){border-color:var(--blue);color:var(--blue)}

/* Tabela Tipo x Matéria */
.tipo-materia-table{width:100%}
.tm-row{margin-bottom:14px;padding-bottom:14px;border-bottom:1px solid var(--border)}
.tm-row:last-child{border-bottom:none;margin-bottom:0;padding-bottom:0}
.tm-materia{font-weight:700;font-size:13px;margin-bottom:6px}
.tm-bars{display:flex;flex-direction:column;gap:4px}
.tm-bar-item{display:grid;grid-template-columns:150px 1fr 40px;gap:6px;align-items:center}
.tm-bar-label{font-size:11px;color:var(--muted)}
.tm-bar-track{background:#f1f5f9;border-radius:3px;height:6px;overflow:hidden}
.tm-bar-fill{height:100%;border-radius:3px}
.tm-bar-val{font-size:11px;font-weight:600;text-align:right}

/* EMPTY STATE */
.empty-state{text-align:center;padding:60px 20px;color:var(--muted)}
.empty-state-icon{font-size:48px;margin-bottom:12px}
.empty-state-text{font-size:15px}

/* RESPONSIVE */
@media(max-width:1024px){
  .kpi-grid{grid-template-columns:1fr 1fr}
  .fase-cards{grid-template-columns:1fr 1fr}
  .grid-2,.grid-3{grid-template-columns:1fr}
  .insight-cards{grid-template-columns:1fr 1fr}
}
@media(max-width:768px){
  .kpi-grid{grid-template-columns:1fr}
  .topbar{flex-wrap:wrap;padding:8px 16px}
  .tabs{flex-wrap:wrap}
  .content{padding:16px}
  .side-panel{width:100%;right:-100%}
  .side-panel.open{right:0}
}

/* OVERLAY */
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.3);z-index:99}
.overlay.show{display:block}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-logo">MEGE<span>.</span> Suporte</div>
  <div class="tabs">
    <button class="tab-btn active" onclick="switchTab(0)">Visão Geral</button>
    <button class="tab-btn" onclick="switchTab(1)">Natureza</button>
    <button class="tab-btn" onclick="switchTab(2)">Por Carreira</button>
    <button class="tab-btn" onclick="switchTab(3)">Explorador</button>
  </div>
</div>

<!-- TAB 0: VISÃO GERAL -->
<div class="content active" id="tab0">
  <div class="filter-bar" id="filter-bar-0">
    <div class="filter-group">
      <label>Carreira</label>
      <select id="f0-carreira" onchange="applyFilters()">
        <option value="">Todas</option>
        <option value="magistratura">Magistratura</option>
        <option value="ministerio_publico">Ministério Público</option>
        <option value="defensoria">Defensoria</option>
        <option value="federal">Federal</option>
        <option value="geral">Geral</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Fase</label>
      <select id="f0-fase" onchange="applyFilters()">
        <option value="">Todas</option>
        <option value="2ª Fase">2ª Fase</option>
        <option value="Pré-Edital">Pré-Edital</option>
        <option value="Clube">Clube</option>
        <option value="Pós-Edital">Pós-Edital</option>
        <option value="Até Passar">Até Passar</option>
        <option value="1ª Fase">1ª Fase</option>
        <option value="Reta Final">Reta Final</option>
        <option value="Geral">Geral</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Turma</label>
      <select id="f0-turma" onchange="applyFilters()">
        <option value="">Todas</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Mês</label>
      <select id="f0-mes" onchange="applyFilters()">
        <option value="">Todos</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Tipo de Suporte</label>
      <select id="f0-tiposuporte" onchange="applyFilters()">
        <option value="">Todos</option>
      </select>
    </div>
    <button class="btn-clear" onclick="clearFilters()">✕ Limpar</button>
  </div>

  <div class="kpi-grid">
    <div class="kpi">
      <div class="kpi-label">Total de Tickets</div>
      <div class="kpi-value" id="kpi-total">—</div>
      <div class="kpi-sub" id="kpi-total-sub">todos os registros</div>
    </div>
    <div class="kpi orange">
      <div class="kpi-label">Resolvidos</div>
      <div class="kpi-value" id="kpi-resolvidos">—</div>
      <div class="kpi-sub" id="kpi-resolvidos-sub">—</div>
    </div>
    <div class="kpi green">
      <div class="kpi-label">Aguardando</div>
      <div class="kpi-value" id="kpi-aguardando">—</div>
      <div class="kpi-sub" id="kpi-aguardando-sub">—</div>
    </div>
    <div class="kpi purple">
      <div class="kpi-label">Tipo Mais Frequente</div>
      <div class="kpi-value" style="font-size:18px" id="kpi-top-tipo">—</div>
      <div class="kpi-sub" id="kpi-top-tipo-sub">—</div>
    </div>
  </div>

  <div class="grid-2">
    <div class="card col-span-2">
      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px">
        <div class="card-title" style="margin-bottom:0">Evolução Temporal</div>
        <div class="toggle-group">
          <button class="toggle-btn active" id="tgl-dia" onclick="setTimeline('dia')">Dia</button>
          <button class="toggle-btn" id="tgl-semana" onclick="setTimeline('semana')">Semana</button>
          <button class="toggle-btn" id="tgl-mes" onclick="setTimeline('mes')">Mês</button>
        </div>
      </div>
      <div class="chart-wrap"><canvas id="chart-timeline" height="80"></canvas></div>
    </div>
  </div>

  <div class="grid-2">
    <div class="card">
      <div class="card-title">Tipos de Solicitação</div>
      <div id="chart-tipos" class="hbar-list"></div>
    </div>
    <div class="card">
      <div class="card-title">Top 10 Turmas <span class="badge-count" id="turma-click-hint" style="font-size:10px;color:var(--blue)">clique para filtrar</span></div>
      <div id="chart-turmas" class="hbar-list"></div>
    </div>
  </div>

  <div class="grid-2">
    <div class="card">
      <div class="card-title">Volume por Dia da Semana</div>
      <div class="chart-wrap"><canvas id="chart-dow" height="100"></canvas></div>
    </div>
    <div class="card">
      <div class="card-title">Material Referenciado</div>
      <div id="chart-material" class="hbar-list"></div>
    </div>
  </div>
</div>

<!-- TAB 1: NATUREZA -->
<div class="content" id="tab1">
  <div class="section-header">
    <div>
      <div class="section-title">Natureza dos Tickets por Fase</div>
      <div class="section-sub">Padrões de problema por fase do curso</div>
    </div>
  </div>

  <div class="fase-cards" id="fase-cards"></div>

  <div class="insight-cards" id="insight-cards"></div>

  <div class="grid-2">
    <div class="card col-span-2">
      <div class="card-title">Heatmap — Matéria × Material Referenciado</div>
      <div class="heatmap-wrap"><div id="heatmap-container"></div></div>
    </div>
  </div>

  <div class="card">
    <div class="card-title">Tipo de Solicitação por Matéria</div>
    <div id="tipo-materia-chart"></div>
  </div>
</div>

<!-- TAB 2: POR CARREIRA -->
<div class="content" id="tab2">
  <div class="section-header">
    <div>
      <div class="section-title">Análise por Carreira</div>
      <div class="section-sub">Comparativo e detalhamento por carreira</div>
    </div>
  </div>

  <div class="carreira-pills" id="carreira-pills">
    <button class="pill active" data-carreira="" onclick="setCarreira(this,'')">Todas</button>
    <button class="pill" data-carreira="magistratura" onclick="setCarreira(this,'magistratura')">Magistratura</button>
    <button class="pill" data-carreira="ministerio_publico" onclick="setCarreira(this,'ministerio_publico')">Ministério Público</button>
    <button class="pill" data-carreira="defensoria" onclick="setCarreira(this,'defensoria')">Defensoria</button>
    <button class="pill" data-carreira="federal" onclick="setCarreira(this,'federal')">Federal</button>
    <button class="pill" data-carreira="geral" onclick="setCarreira(this,'geral')">Geral</button>
  </div>

  <div class="grid-2">
    <div class="card">
      <div class="card-title">Tipos de Solicitação por Carreira</div>
      <div class="chart-wrap"><canvas id="chart-carreira-tipo" height="180"></canvas></div>
    </div>
    <div class="card">
      <div class="card-title">Matéria por Carreira</div>
      <div class="chart-wrap"><canvas id="chart-carreira-materia" height="180"></canvas></div>
    </div>
  </div>

  <div class="grid-2">
    <div class="card">
      <div class="card-title">Backlog (Aguardando) por Carreira</div>
      <div id="backlog-carreira" class="hbar-list"></div>
    </div>
    <div class="card">
      <div class="card-title" id="carreira-insights-title">Insights</div>
      <div id="carreira-insights" style="display:flex;flex-direction:column;gap:10px"></div>
    </div>
  </div>
</div>

<!-- TAB 3: EXPLORADOR -->
<div class="content" id="tab3">
  <div class="filter-bar">
    <div class="filter-group" style="flex:1;min-width:200px">
      <label>Buscar no título</label>
      <input type="text" id="f3-busca" placeholder="Digite para buscar..." oninput="renderExplorer()"/>
    </div>
    <div class="filter-group">
      <label>Tipo</label>
      <select id="f3-tipo" onchange="renderExplorer()">
        <option value="">Todos</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Carreira</label>
      <select id="f3-carreira" onchange="renderExplorer()">
        <option value="">Todas</option>
        <option value="magistratura">Magistratura</option>
        <option value="ministerio_publico">Ministério Público</option>
        <option value="defensoria">Defensoria</option>
        <option value="federal">Federal</option>
        <option value="geral">Geral</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Fase</label>
      <select id="f3-fase" onchange="renderExplorer()">
        <option value="">Todas</option>
        <option value="2ª Fase">2ª Fase</option>
        <option value="Pré-Edital">Pré-Edital</option>
        <option value="Clube">Clube</option>
        <option value="Pós-Edital">Pós-Edital</option>
        <option value="Até Passar">Até Passar</option>
        <option value="1ª Fase">1ª Fase</option>
        <option value="Reta Final">Reta Final</option>
        <option value="Geral">Geral</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Matéria</label>
      <select id="f3-materia" onchange="renderExplorer()">
        <option value="">Todas</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Status</label>
      <select id="f3-status" onchange="renderExplorer()">
        <option value="">Todos</option>
        <option value="Resolvido">Resolvido</option>
        <option value="Aguardando">Aguardando</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Tipo de Suporte</label>
      <select id="f3-tiposuporte" onchange="renderExplorer()">
        <option value="">Todos</option>
      </select>
    </div>
  </div>

  <div class="result-count" id="result-count">Carregando...</div>

  <div class="tbl-wrap">
    <table class="tbl">
      <thead>
        <tr>
          <th>ID</th>
          <th>Data</th>
          <th>Título</th>
          <th>Tipo</th>
          <th>Carreira</th>
          <th>Fase</th>
          <th>Matéria</th>
          <th>Status</th>
          <th>Conf.</th>
        </tr>
      </thead>
      <tbody id="explorer-tbody"></tbody>
    </table>
  </div>
  <div class="pagination" id="pagination"></div>
</div>

<!-- SIDE PANEL -->
<div class="overlay" id="overlay" onclick="closePanel()"></div>
<div class="side-panel" id="side-panel">
  <div class="side-panel-header">
    <div class="side-panel-titulo" id="sp-titulo">—</div>
    <button class="btn-close" onclick="closePanel()">✕</button>
  </div>
  <div class="side-panel-body">
    <div id="sp-id-box" style="display:none;align-items:center;gap:8px;background:#f1f5f9;border-radius:8px;padding:8px 12px;margin-bottom:14px;font-size:11px;font-family:monospace;color:var(--muted)">
      <span style="font-weight:600;color:var(--text);font-family:'Inter',sans-serif;font-size:10px;text-transform:uppercase;letter-spacing:.05em;flex-shrink:0">ID</span>
      <span id="sp-id-val" style="flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"></span>
      <button id="sp-id-copy" title="Copiar ID" style="background:none;border:1px solid var(--border);border-radius:5px;padding:2px 7px;cursor:pointer;font-size:13px;flex-shrink:0;color:var(--muted)">⧉</button>
    </div>
    <div class="sp-badges" id="sp-badges"></div>
    <div class="sp-meta" id="sp-meta"></div>
    <div class="sp-section">
      <div class="sp-section-label">Descrição do Aluno</div>
      <div class="sp-box" id="sp-desc">—</div>
    </div>
    <div class="sp-section">
      <div class="sp-section-label">Resposta do Coordenador</div>
      <div class="sp-box green" id="sp-resp">—</div>
    </div>
    <div class="sp-section" id="sp-nota-sec" style="display:none">
      <div class="sp-section-label">Nota de Ambiguidade</div>
      <div class="sp-box yellow" id="sp-nota">—</div>
    </div>
  </div>
</div>

<script>
// ============================================================
// 1. DATA
// ============================================================
// Filter array columns: [0]=data [1]=tipo [2]=carreira [3]=fase
// [4]=materia [5]=material [6]=turma [7]=status [8]=mes [9]=semana [10]=dow
const FD = __FILTER_DATA__;

// Full array columns: [0]=data [1]=titulo [2]=tipo [3]=carreira [4]=fase
// [5]=materia [6]=material [7]=turma [8]=status [9]=coord [10]=conf
// [11]=desc [12]=resp [13]=nota [14]=mes
const DD = __FULL_DATA__;

// ============================================================
// 2. CONSTANTS & COLORS
// ============================================================
const TIPO_COLORS = {
  extensao_prazo:'#09408C', informacao_geral:'#F58220', correcao_avaliacao:'#029A42',
  material_ausente:'#6400FC', duvida_juridica:'#0694A2', administrativo:'#D53F8C',
  erro_material:'#E53E3E', liberacao_envio:'#D69E2E', acesso_tecnico:'#2D3748', outro:'#718096'
};
const TIPO_LABELS = {
  extensao_prazo:'Extensão de Prazo', informacao_geral:'Informação Geral',
  correcao_avaliacao:'Correção de Avaliação', material_ausente:'Material Ausente',
  duvida_juridica:'Dúvida Jurídica', administrativo:'Administrativo',
  erro_material:'Erro de Material', liberacao_envio:'Liberação de Envio',
  acesso_tecnico:'Acesso Técnico', outro:'Outro'
};
const CARREIRA_LABELS = {
  magistratura:'Magistratura', ministerio_publico:'Ministério Público',
  defensoria:'Defensoria', federal:'Federal', geral:'Geral'
};
const CARREIRA_COLORS = {
  magistratura:'#09408C', ministerio_publico:'#F58220', defensoria:'#029A42',
  federal:'#6400FC', geral:'#718096'
};
const MATERIA_LABELS = {
  processo_civil:'Proc. Civil', processo_penal:'Proc. Penal', direito_penal:'Dir. Penal',
  direito_civil:'Dir. Civil', constitucional:'Constitucional', planejamento_estudos:'Planejamento',
  humanistica:'Humanística', administrativo:'Administrativo', nao_identificada:'Não Identificada',
  ambiental:'Ambiental', consumidor:'Consumidor', empresarial:'Empresarial',
  informativos_juris:'Informativos'
};
const MATERIAL_LABELS = {
  sentenca_discursiva:'Sentença Discursiva', rodada_ponto_ebook:'Rodada/Ebook',
  simulado_objetivo:'Simulado Objetivo', videoaula:'Videoaula',
  circuito_legislativo:'Circuito Leg.', vade_mege:'Vade MEGE',
  informativos:'Informativos', nenhum:'Nenhum'
};
const FASE_COLORS = {
  '2ª Fase':'#09408C', 'Pré-Edital':'#F58220', 'Clube':'#029A42',
  'Pós-Edital':'#6400FC', 'Até Passar':'#D53F8C', '1ª Fase':'#0694A2',
  'Reta Final':'#D69E2E', 'Geral':'#718096'
};
const DOW_LABELS = ['Seg','Ter','Qua','Qui','Sex','Sáb','Dom'];
const TIPOSUPORTE_LABELS = {
  'Conteúdo / Material de Apoio (Observaçōes) ': 'Conteúdo / Material de Apoio',
  'Conteúdo / Material de Apoio (Observações) ': 'Conteúdo / Material de Apoio',
  'Coordenação Acadêmica/Institucional': 'Coordenação Acadêmica',
  'Assuntos Administrativos': 'Administrativo',
  'Simulados de provas Objetivas (Questōes)': 'Simulados Objetivos',
  'Simulados de provas Objetivas (Questões)': 'Simulados Objetivos',
  'Perguntas das avaliaçōes': 'Perguntas de Avaliações',
  'Perguntas das avaliações': 'Perguntas de Avaliações',
  'Financeiro / Atendimento': 'Financeiro / Atendimento',
  'Mege Informativos (Aulas e materiais de Jurisprudência)': 'MEGE Informativos',
};

function hl(key, map){ return map[key] || key.replace(/_/g,' ').replace(/\b\w/g,c=>c.toUpperCase()); }

// ============================================================
// 3. STATE
// ============================================================
let activeCarreira = '';
let timelineMode = 'dia';
let explorerPage = 0;
const PAGE_SIZE = 30;
let explorerFiltered = [];
let selectedRow = -1;

// Chart instances
let chartTimeline=null, chartDow=null, chartCarreiraTipo=null, chartCarreiraMateria=null;

// ============================================================
// 4. FILTERS (Tab 0)
// ============================================================
function getFiltered(){
  const carreira = document.getElementById('f0-carreira').value;
  const fase = document.getElementById('f0-fase').value;
  const turma = document.getElementById('f0-turma').value;
  const mes = document.getElementById('f0-mes').value;
  const tiposuporte = document.getElementById('f0-tiposuporte').value;
  if(!carreira && !fase && !turma && !mes && !tiposuporte) return FD;
  return FD.filter(r => {
    if(carreira && r[2]!==carreira) return false;
    if(fase && r[3]!==fase) return false;
    if(turma && r[6]!==turma) return false;
    if(mes && r[8]!==mes) return false;
    if(tiposuporte && r[11]!==tiposuporte) return false;
    return true;
  });
}

function applyFilters(){
  const data = getFiltered();
  updateKPIs(data);
  updateTimeline(data);
  updateTipos(data);
  updateTurmas(data);
  updateDow(data);
  updateMaterial(data);
}

function clearFilters(){
  document.getElementById('f0-carreira').value='';
  document.getElementById('f0-fase').value='';
  document.getElementById('f0-turma').value='';
  document.getElementById('f0-mes').value='';
  document.getElementById('f0-tiposuporte').value='';
  applyFilters();
}

function setTurmaFilter(turma){
  document.getElementById('f0-turma').value=turma;
  applyFilters();
}

// ============================================================
// 5. KPIs
// ============================================================
function updateKPIs(data){
  const total = data.length;
  const resolvidos = data.filter(r=>r[7]==='Resolvido').length;
  const aguardando = total - resolvidos;
  document.getElementById('kpi-total').textContent = total.toLocaleString('pt-BR');
  document.getElementById('kpi-resolvidos').textContent = resolvidos.toLocaleString('pt-BR');
  document.getElementById('kpi-resolvidos-sub').textContent = total?((resolvidos/total*100).toFixed(1)+'% do total'):'—';
  document.getElementById('kpi-aguardando').textContent = aguardando.toLocaleString('pt-BR');
  document.getElementById('kpi-aguardando-sub').textContent = total?((aguardando/total*100).toFixed(1)+'% em aberto'):'—';
  // Top tipo
  const tipoCount = count(data,1);
  const topTipo = tipoCount.length ? tipoCount[0] : null;
  document.getElementById('kpi-top-tipo').textContent = topTipo ? hl(topTipo[0],TIPO_LABELS) : '—';
  document.getElementById('kpi-top-tipo-sub').textContent = topTipo ? (topTipo[1].toLocaleString('pt-BR')+' tickets') : '—';
}

// ============================================================
// 6. HELPERS
// ============================================================
function count(data, col){
  const m={};
  data.forEach(r=>{ const k=r[col]||''; m[k]=(m[k]||0)+1; });
  return Object.entries(m).sort((a,b)=>b[1]-a[1]);
}

function renderHBars(container, items, total, colorMap, labelMap, clickFn, maxItems){
  if(maxItems) items = items.slice(0,maxItems);
  const maxVal = items.length ? items[0][1] : 1;
  let html='<div class="hbar-list">';
  // Calculate pareto
  let cumSum=0, paretoIdx=-1;
  const totalForPareto = items.reduce((s,x)=>s+x[1],0);
  for(let i=0;i<items.length;i++){
    cumSum+=items[i][1];
    if(cumSum/totalForPareto>=0.8 && paretoIdx===-1){ paretoIdx=i; }
  }
  items.forEach(([k,v],i)=>{
    const pct = total?((v/total)*100).toFixed(1):0;
    const barPct = maxVal?((v/maxVal)*100).toFixed(1):0;
    const color = (colorMap&&colorMap[k])||'#09408C';
    const label = labelMap ? hl(k,labelMap) : k;
    const clickAttr = clickFn ? `onclick="${clickFn}('${k.replace(/'/g,"\\'")}')"`:'';
    const clickable = clickFn ? 'clickable':'';
    html+=`<div class="hbar-item ${clickable}" ${clickAttr}>
      <span class="hbar-label" title="${label}">${label}</span>
      <div class="hbar-track"><div class="hbar-fill" style="width:${barPct}%;background:${color}"></div></div>
      <span class="hbar-val">${v.toLocaleString('pt-BR')}</span>
      <span class="hbar-pct">${pct}%</span>
    </div>`;
    if(i===paretoIdx && paretoIdx<items.length-1){
      html+='<div class="pareto-line"></div>';
    }
  });
  html+='</div>';
  document.getElementById(container).innerHTML=html;
}

// ============================================================
// 7. TIMELINE
// ============================================================
function setTimeline(mode){
  timelineMode = mode;
  ['dia','semana','mes'].forEach(m=>{
    document.getElementById('tgl-'+m).classList.toggle('active',m===mode);
  });
  updateTimeline(getFiltered());
}

function updateTimeline(data){
  const col = timelineMode==='dia'?0 : timelineMode==='semana'?9 : 8;
  const m={};
  data.forEach(r=>{ const k=r[col]||''; if(k) m[k]=(m[k]||0)+1; });
  const entries = Object.entries(m).sort((a,b)=>a[0]>b[0]?1:-1);
  const labels = entries.map(e=>formatDateLabel(e[0],timelineMode));
  const values = entries.map(e=>e[1]);

  if(chartTimeline){ chartTimeline.destroy(); chartTimeline=null; }
  const ctx = document.getElementById('chart-timeline').getContext('2d');
  chartTimeline = new Chart(ctx,{
    type:'bar',
    data:{ labels, datasets:[{ data:values, backgroundColor:'rgba(9,64,140,0.75)', borderRadius:3, borderSkipped:false }] },
    options:{
      responsive:true, maintainAspectRatio:true,
      plugins:{ legend:{display:false}, tooltip:{ backgroundColor:'#fff', borderColor:'#e2e8f0', borderWidth:1, titleColor:'#1a202c', bodyColor:'#1a202c', padding:10 }},
      scales:{ x:{ grid:{display:false}, ticks:{maxTicksLimit:20,maxRotation:45,font:{size:10}} }, y:{ grid:{color:'#f1f5f9'}, ticks:{font:{size:10}} } }
    }
  });
}

function formatDateLabel(key,mode){
  if(!key) return key;
  if(mode==='dia'){ const p=key.split('-'); return p[2]+'/'+p[1]; }
  if(mode==='mes'){ const p=key.split('-'); const ms=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']; return (ms[parseInt(p[1])-1]||p[1])+'/'+p[0].slice(2); }
  return key;
}

// ============================================================
// 8. TIPOS
// ============================================================
function updateTipos(data){
  const items = count(data,1);
  renderHBars('chart-tipos', items, data.length, TIPO_COLORS, TIPO_LABELS, null, null);
}

// ============================================================
// 9. TURMAS
// ============================================================
function updateTurmas(data){
  const items = count(data,6).slice(0,10);
  const total = data.length;
  const maxVal = items.length ? items[0][1] : 1;
  let html='<div class="hbar-list">';
  items.forEach(([k,v])=>{
    const pct = total?((v/total)*100).toFixed(1):0;
    const barPct = maxVal?((v/maxVal)*100).toFixed(1):0;
    const label = k.length>35?k.slice(0,35)+'…':k;
    html+=`<div class="hbar-item clickable" onclick="setTurmaFilter('${k.replace(/'/g,"\\'").replace(/\\/g,"\\\\")}')" >
      <span class="hbar-label" title="${k}">${label}</span>
      <div class="hbar-track"><div class="hbar-fill" style="width:${barPct}%;background:#09408C"></div></div>
      <span class="hbar-val">${v.toLocaleString('pt-BR')}</span>
      <span class="hbar-pct">${pct}%</span>
    </div>`;
  });
  html+='</div>';
  document.getElementById('chart-turmas').innerHTML=html;
}

// ============================================================
// 10. DOW
// ============================================================
function updateDow(data){
  const counts=[0,0,0,0,0,0,0];
  data.forEach(r=>{ if(r[10]>=0 && r[10]<=6) counts[r[10]]++; });
  if(chartDow){ chartDow.destroy(); chartDow=null; }
  const ctx=document.getElementById('chart-dow').getContext('2d');
  const colors = counts.map((_,i)=>i>=5?'rgba(245,130,32,0.75)':'rgba(9,64,140,0.75)');
  chartDow = new Chart(ctx,{
    type:'bar',
    data:{ labels:DOW_LABELS, datasets:[{ data:counts, backgroundColor:colors, borderRadius:4, borderSkipped:false }]},
    options:{
      responsive:true, maintainAspectRatio:true,
      plugins:{ legend:{display:false}, tooltip:{ backgroundColor:'#fff', borderColor:'#e2e8f0', borderWidth:1, titleColor:'#1a202c', bodyColor:'#1a202c' }},
      scales:{ x:{ grid:{display:false}, ticks:{font:{size:11}} }, y:{ grid:{color:'#f1f5f9'}, ticks:{font:{size:10}} }}
    }
  });
}

// ============================================================
// 11. MATERIAL
// ============================================================
function updateMaterial(data){
  const items = count(data,5).filter(([k])=>k!=='nenhum');
  renderHBars('chart-material', items, data.length, null, MATERIAL_LABELS, null, null);
}

// ============================================================
// 12. TAB 1: NATUREZA
// ============================================================
function renderNatureza(){
  const fases = ['2ª Fase','Pré-Edital','Clube','Pós-Edital','Até Passar'];
  const faseData = {};
  fases.forEach(f=>{ faseData[f]=[]; });
  FD.forEach(r=>{ if(faseData[r[3]]) faseData[r[3]].push(r); });

  // Fase cards
  let html='';
  fases.forEach(f=>{
    const d = faseData[f]||[];
    const tipoC = count(d,1).slice(0,3);
    const matC = count(d,5).filter(([k])=>k!=='nenhum');
    const topMat = matC.length ? hl(matC[0][0],MATERIAL_LABELS) : '—';
    const total = d.length;
    const color = FASE_COLORS[f]||'#09408C';
    html+=`<div class="fase-card">
      <div class="fase-card-name" style="color:${color}">${f}</div>
      <div class="fase-card-vol">${total.toLocaleString('pt-BR')}</div>`;
    const maxV = tipoC.length?tipoC[0][1]:1;
    tipoC.forEach(([k,v])=>{
      const pct = total?((v/total)*100).toFixed(0):0;
      const w = maxV?((v/maxV)*100).toFixed(0):0;
      html+=`<div class="fase-mini-bar">
        <div class="fase-mini-label"><span>${hl(k,TIPO_LABELS)}</span><span>${pct}%</span></div>
        <div class="fase-mini-track"><div class="fase-mini-fill" style="width:${w}%;background:${TIPO_COLORS[k]||color}"></div></div>
      </div>`;
    });
    html+=`<div class="fase-material">📎 ${topMat}</div></div>`;
  });
  document.getElementById('fase-cards').innerHTML=html;

  // Insights
  renderInsights(faseData);

  // Heatmap
  renderHeatmap();

  // Tipo x Matéria
  renderTipoMateria();
}

function renderInsights(faseData){
  const allFases = Object.entries(faseData);
  const insights=[];

  // Insight 1: Fase com mais tickets
  const topFase = allFases.sort((a,b)=>b[1].length-a[1].length)[0];
  insights.push({icon:'📊', title:`${topFase[0]} lidera em volume`, text:`Com ${topFase[1].length.toLocaleString('pt-BR')} tickets, é a fase com maior demanda de suporte.`});

  // Insight 2: Tipo dominante
  const tipoC = count(FD,1);
  if(tipoC.length){
    const [k,v] = tipoC[0];
    insights.push({icon:'🔝', title:`${hl(k,TIPO_LABELS)} é o tipo mais comum`, text:`Representa ${((v/FD.length)*100).toFixed(1)}% de todos os tickets. Foco neste tipo pode reduzir volume significativamente.`});
  }

  // Insight 3: Taxa de resolução
  const resolvidos = FD.filter(r=>r[7]==='Resolvido').length;
  const pct = ((resolvidos/FD.length)*100).toFixed(1);
  insights.push({icon:'✅', title:`Taxa de resolução de ${pct}%`, text:`${resolvidos.toLocaleString('pt-BR')} tickets resolvidos de ${FD.length.toLocaleString('pt-BR')} total.`});

  // Insight 4: Material mais referenciado
  const matC = count(FD,5).filter(([k])=>k!=='nenhum');
  if(matC.length){
    const [k,v] = matC[0];
    insights.push({icon:'📚', title:`${hl(k,MATERIAL_LABELS)} gera mais dúvidas`, text:`Aparece em ${v.toLocaleString('pt-BR')} tickets, sendo o material mais referenciado.`});
  }

  let html='';
  insights.forEach(({icon,title,text})=>{
    html+=`<div class="insight-card">
      <div class="insight-icon">${icon}</div>
      <div class="insight-title">${title}</div>
      <div class="insight-text">${text}</div>
    </div>`;
  });
  document.getElementById('insight-cards').innerHTML=html;
}

function renderHeatmap(){
  const materias = [...new Set(FD.map(r=>r[4]).filter(x=>x&&x!=='nao_identificada'))].sort();
  const materiais = [...new Set(FD.map(r=>r[5]).filter(x=>x&&x!=='nenhum'))].sort();

  // Build matrix
  const m={};
  FD.forEach(r=>{
    if(!r[4]||r[4]==='nao_identificada') return;
    if(!r[5]||r[5]==='nenhum') return;
    const k=r[4]+'|'+r[5];
    m[k]=(m[k]||0)+1;
  });
  const vals = Object.values(m);
  const maxV = vals.length ? Math.max(...vals) : 1;

  let html='<table class="heatmap"><thead><tr><th>Matéria \\ Material</th>';
  materiais.forEach(mat=>{ html+=`<th>${hl(mat,MATERIAL_LABELS)}</th>`; });
  html+='</tr></thead><tbody>';
  materias.forEach(mat=>{
    html+=`<tr><td class="row-label">${hl(mat,MATERIA_LABELS)}</td>`;
    materiais.forEach(material=>{
      const v = m[mat+'|'+material]||0;
      const intensity = maxV ? v/maxV : 0;
      const alpha = (0.1 + intensity*0.85).toFixed(2);
      const textColor = intensity>0.5 ? '#fff':'#1a202c';
      html+=`<td style="background:rgba(9,64,140,${alpha});color:${textColor};font-weight:${v?600:400}">${v||''}</td>`;
    });
    html+='</tr>';
  });
  html+='</tbody></table>';
  document.getElementById('heatmap-container').innerHTML=html;
}

function renderTipoMateria(){
  const materias = [...new Set(FD.map(r=>r[4]).filter(x=>x&&x!=='nao_identificada'))];
  const mData = {};
  materias.forEach(m=>{ mData[m]={}; });
  FD.forEach(r=>{
    if(!r[4]||r[4]==='nao_identificada') return;
    mData[r[4]][r[1]]=(mData[r[4]][r[1]]||0)+1;
  });

  // Sort materias by total
  const matSorted = materias.map(m=>([m, Object.values(mData[m]).reduce((s,v)=>s+v,0)])).sort((a,b)=>b[1]-a[1]);

  let html='<div class="tipo-materia-table">';
  matSorted.forEach(([mat,total])=>{
    const tipos = Object.entries(mData[mat]).sort((a,b)=>b[1]-a[1]).slice(0,3);
    const maxV = tipos.length?tipos[0][1]:1;
    html+=`<div class="tm-row">
      <div class="tm-materia">${hl(mat,MATERIA_LABELS)} <span style="font-weight:400;color:var(--muted);font-size:12px">(${total.toLocaleString('pt-BR')})</span></div>
      <div class="tm-bars">`;
    tipos.forEach(([k,v])=>{
      const w = maxV?((v/maxV)*100).toFixed(0):0;
      const color = TIPO_COLORS[k]||'#09408C';
      html+=`<div class="tm-bar-item">
        <span class="tm-bar-label">${hl(k,TIPO_LABELS)}</span>
        <div class="tm-bar-track"><div class="tm-bar-fill" style="width:${w}%;background:${color}"></div></div>
        <span class="tm-bar-val">${v}</span>
      </div>`;
    });
    html+='</div></div>';
  });
  html+='</div>';
  document.getElementById('tipo-materia-chart').innerHTML=html;
}

// ============================================================
// 13. TAB 2: POR CARREIRA
// ============================================================
function setCarreira(el,carreira){
  activeCarreira = carreira;
  document.querySelectorAll('.carreira-pills .pill').forEach(p=>p.classList.remove('active'));
  el.classList.add('active');
  // Color the pill
  const color = carreira ? CARREIRA_COLORS[carreira]||'#09408C' : '#09408C';
  if(carreira) el.style.background=color;
  document.querySelectorAll('.carreira-pills .pill').forEach(p=>{
    if(!p.classList.contains('active')) p.style.background='';
  });
  renderCarreira();
}

function renderCarreira(){
  const data = activeCarreira ? FD.filter(r=>r[2]===activeCarreira) : FD;
  const carreiras = activeCarreira ? [activeCarreira] : ['magistratura','ministerio_publico','defensoria','federal','geral'];

  // Chart: tipo por carreira
  const tipos = Object.keys(TIPO_LABELS);
  const datasets = carreiras.map(c=>{
    const d = FD.filter(r=>r[2]===c);
    const tc = {};
    d.forEach(r=>{ tc[r[1]]=(tc[r[1]]||0)+1; });
    return {
      label: hl(c,CARREIRA_LABELS),
      data: tipos.map(t=>tc[t]||0),
      backgroundColor: CARREIRA_COLORS[c]||'#718096',
      borderRadius:3
    };
  });

  if(chartCarreiraTipo){ chartCarreiraTipo.destroy(); chartCarreiraTipo=null; }
  const ctx1 = document.getElementById('chart-carreira-tipo').getContext('2d');
  chartCarreiraTipo = new Chart(ctx1,{
    type:'bar',
    data:{ labels: tipos.map(t=>hl(t,TIPO_LABELS)), datasets },
    options:{
      responsive:true, maintainAspectRatio:true, indexAxis:'y',
      plugins:{ legend:{display:carreiras.length>1, position:'top', labels:{font:{size:11},boxWidth:12} },
        tooltip:{ backgroundColor:'#fff', borderColor:'#e2e8f0', borderWidth:1, titleColor:'#1a202c', bodyColor:'#1a202c' }},
      scales:{
        x:{ stacked:false, grid:{color:'#f1f5f9'}, ticks:{font:{size:10}} },
        y:{ grid:{display:false}, ticks:{font:{size:10}} }
      }
    }
  });

  // Chart: materia por carreira
  const materias = [...new Set(FD.map(r=>r[4]).filter(x=>x))].sort();
  const datasets2 = carreiras.map(c=>{
    const d = FD.filter(r=>r[2]===c);
    const mc={};
    d.forEach(r=>{ mc[r[4]]=(mc[r[4]]||0)+1; });
    return {
      label:hl(c,CARREIRA_LABELS),
      data:materias.map(m=>mc[m]||0),
      backgroundColor:CARREIRA_COLORS[c]||'#718096',
      borderRadius:3
    };
  });

  if(chartCarreiraMateria){ chartCarreiraMateria.destroy(); chartCarreiraMateria=null; }
  const ctx2 = document.getElementById('chart-carreira-materia').getContext('2d');
  chartCarreiraMateria = new Chart(ctx2,{
    type:'bar',
    data:{ labels:materias.map(m=>hl(m,MATERIA_LABELS)), datasets:datasets2 },
    options:{
      responsive:true, maintainAspectRatio:true, indexAxis:'y',
      plugins:{ legend:{display:carreiras.length>1, position:'top', labels:{font:{size:11},boxWidth:12} },
        tooltip:{ backgroundColor:'#fff', borderColor:'#e2e8f0', borderWidth:1, titleColor:'#1a202c', bodyColor:'#1a202c' }},
      scales:{
        x:{stacked:false, grid:{color:'#f1f5f9'}, ticks:{font:{size:10}} },
        y:{ grid:{display:false}, ticks:{font:{size:10}} }
      }
    }
  });

  // Backlog
  const backlogItems = (['magistratura','ministerio_publico','defensoria','federal','geral']).map(c=>{
    const total = FD.filter(r=>r[2]===c).length;
    const ag = FD.filter(r=>r[2]===c&&r[7]==='Aguardando').length;
    return [hl(c,CARREIRA_LABELS), ag, total];
  }).sort((a,b)=>b[1]-a[1]);
  const maxBL = backlogItems.length?backlogItems[0][1]:1;
  let blHtml='<div class="hbar-list">';
  backlogItems.forEach(([label,ag,total])=>{
    const w = maxBL?((ag/maxBL)*100).toFixed(0):0;
    const pct = total?((ag/total)*100).toFixed(1):0;
    blHtml+=`<div class="hbar-item">
      <span class="hbar-label">${label}</span>
      <div class="hbar-track"><div class="hbar-fill" style="width:${w}%;background:#E53E3E"></div></div>
      <span class="hbar-val">${ag}</span>
      <span class="hbar-pct">${pct}%</span>
    </div>`;
  });
  blHtml+='</div>';
  document.getElementById('backlog-carreira').innerHTML=blHtml;

  // Insights
  renderCarreiraInsights(data, activeCarreira);
}

function renderCarreiraInsights(data, carreira){
  const total = data.length;
  const resolvidos = data.filter(r=>r[7]==='Resolvido').length;
  const aguardando = total - resolvidos;
  const tipoC = count(data,1);
  const topTipo = tipoC.length?tipoC[0]:null;
  const matC = count(data,4);
  const topMat = matC.length?matC[0]:null;
  const materia2 = count(data,5).filter(([k])=>k!=='nenhum');
  const topMaterial = materia2.length?materia2[0]:null;

  const title = carreira ? hl(carreira,CARREIRA_LABELS) : 'Geral';
  document.getElementById('carreira-insights-title').textContent = `Insights — ${title}`;

  const items=[
    {icon:'📋', text:`<strong>${total.toLocaleString('pt-BR')}</strong> tickets no total`},
    {icon:'✅', text:`Taxa de resolução: <strong>${total?((resolvidos/total)*100).toFixed(1):0}%</strong> (${resolvidos.toLocaleString('pt-BR')} resolvidos)`},
    {icon:'⏳', text:`Em aberto: <strong>${aguardando.toLocaleString('pt-BR')}</strong> tickets aguardando`},
    topTipo?{icon:'🔝', text:`Tipo mais frequente: <strong>${hl(topTipo[0],TIPO_LABELS)}</strong> (${topTipo[1]} tickets)`}:null,
    topMat?{icon:'📖', text:`Matéria mais frequente: <strong>${hl(topMat[0],MATERIA_LABELS)}</strong> (${topMat[1]} tickets)`}:null,
    topMaterial?{icon:'📎', text:`Material mais referenciado: <strong>${hl(topMaterial[0],MATERIAL_LABELS)}</strong> (${topMaterial[1]} tickets)`}:null,
  ].filter(Boolean);

  document.getElementById('carreira-insights').innerHTML = items.map(({icon,text})=>
    `<div style="display:flex;gap:10px;align-items:flex-start;padding:10px;background:var(--bg);border-radius:8px">
      <span style="font-size:18px">${icon}</span>
      <span style="font-size:13px;color:var(--text)">${text}</span>
    </div>`
  ).join('');
}

// ============================================================
// 14. TAB 3: EXPLORADOR
// ============================================================
function buildBadge(text, color){
  return `<span class="badge" style="background:${color}20;color:${color};border:1px solid ${color}40">${text}</span>`;
}

function confBadge(conf){
  if(conf===null||conf===undefined) return '—';
  const c = conf>=0.9?'#029A42':conf>=0.7?'#D69E2E':'#E53E3E';
  return `<span class="badge" style="background:${c}20;color:${c};border:1px solid ${c}40">${(conf*100).toFixed(0)}%</span>`;
}

function statusBadge(s){
  const c = s==='Resolvido'?'#029A42':'#F58220';
  return `<span class="badge" style="background:${c}20;color:${c};border:1px solid ${c}40">${s}</span>`;
}

function renderExplorer(){
  const busca = document.getElementById('f3-busca').value.toLowerCase();
  const tipo = document.getElementById('f3-tipo').value;
  const carreira = document.getElementById('f3-carreira').value;
  const fase = document.getElementById('f3-fase').value;
  const materia = document.getElementById('f3-materia').value;
  const status = document.getElementById('f3-status').value;
  const tiposuporte = document.getElementById('f3-tiposuporte').value;

  explorerFiltered = DD.filter(r=>{
    if(busca && !r[1].toLowerCase().includes(busca)) return false;
    if(tipo && r[2]!==tipo) return false;
    if(carreira && r[3]!==carreira) return false;
    if(fase && r[4]!==fase) return false;
    if(materia && r[5]!==materia) return false;
    if(status && r[8]!==status) return false;
    if(tiposuporte && r[15]!==tiposuporte) return false;
    return true;
  });

  explorerPage=0;
  renderExplorerPage();
}

function renderExplorerPage(){
  const total = explorerFiltered.length;
  const start = explorerPage*PAGE_SIZE;
  const end = Math.min(start+PAGE_SIZE, total);
  const page = explorerFiltered.slice(start,end);

  document.getElementById('result-count').innerHTML=`<strong>${total.toLocaleString('pt-BR')}</strong> tickets encontrados`;

  let html='';
  if(page.length===0){
    html=`<tr><td colspan="8" style="text-align:center;padding:40px;color:var(--muted)">Nenhum ticket encontrado</td></tr>`;
  } else {
    page.forEach((r,i)=>{
      const idx = start+i;
      const tipo = r[2], carreira=r[3], fase=r[4], status=r[8], conf=r[10];
      const tColor = TIPO_COLORS[tipo]||'#718096';
      const cColor = CARREIRA_COLORS[carreira]||'#718096';
      const shortId = r[16] ? r[16].slice(0,8)+'…' : '—';
      html+=`<tr onclick="openPanel(${idx})" class="${selectedRow===idx?'selected':''}">
        <td style="white-space:nowrap;font-size:11px;color:var(--muted);font-family:monospace" title="${r[16]||''}">${shortId}</td>
        <td style="white-space:nowrap;color:var(--muted);font-size:12px">${r[0]?r[0].slice(5):''}</td>
        <td class="tbl-titulo" title="${r[1].replace(/"/g,'&quot;')}">${r[1]}</td>
        <td>${buildBadge(hl(tipo,TIPO_LABELS),tColor)}</td>
        <td>${buildBadge(hl(carreira,CARREIRA_LABELS),cColor)}</td>
        <td style="font-size:12px;white-space:nowrap">${fase}</td>
        <td style="font-size:12px">${hl(r[5],MATERIA_LABELS)}</td>
        <td>${statusBadge(status)}</td>
        <td>${confBadge(conf)}</td>
      </tr>`;
    });
  }
  document.getElementById('explorer-tbody').innerHTML=html;
  renderPagination(total);
}

function renderPagination(total){
  const totalPages = Math.ceil(total/PAGE_SIZE);
  if(totalPages<=1){ document.getElementById('pagination').innerHTML=''; return; }
  let html=`<button class="pg-btn" onclick="goPage(${explorerPage-1})" ${explorerPage===0?'disabled':''}>←</button>`;
  const start=Math.max(0,explorerPage-2), end=Math.min(totalPages,explorerPage+3);
  if(start>0) html+=`<button class="pg-btn" onclick="goPage(0)">1</button>${start>1?'<span class="pg-info">…</span>':''}`;
  for(let p=start;p<end;p++){
    html+=`<button class="pg-btn ${p===explorerPage?'active':''}" onclick="goPage(${p})">${p+1}</button>`;
  }
  if(end<totalPages) html+=`${end<totalPages-1?'<span class="pg-info">…</span>':''}<button class="pg-btn" onclick="goPage(${totalPages-1})">${totalPages}</button>`;
  html+=`<button class="pg-btn" onclick="goPage(${explorerPage+1})" ${explorerPage===totalPages-1?'disabled':''}>→</button>`;
  html+=`<span class="pg-info">${explorerPage*PAGE_SIZE+1}–${Math.min((explorerPage+1)*PAGE_SIZE,total)} de ${total.toLocaleString('pt-BR')}</span>`;
  document.getElementById('pagination').innerHTML=html;
}

function goPage(p){
  explorerPage=p;
  renderExplorerPage();
  document.getElementById('tab3').scrollIntoView({behavior:'smooth',block:'start'});
}

function openPanel(idx){
  selectedRow = idx;
  const r = explorerFiltered[idx];
  if(!r) return;

  document.getElementById('sp-titulo').textContent = r[1]||'Sem título';
  const idFull = r[16]||'';
  document.getElementById('sp-id-box').style.display = idFull ? 'flex' : 'none';
  document.getElementById('sp-id-val').textContent = idFull;
  document.getElementById('sp-id-copy').onclick = ()=>{
    navigator.clipboard.writeText(idFull).then(()=>{
      const btn = document.getElementById('sp-id-copy');
      btn.textContent='✓'; setTimeout(()=>btn.textContent='⧉',1500);
    });
  };

  const tipo=r[2],carreira=r[3],fase=r[4],materia=r[5],material=r[6];
  const tColor=TIPO_COLORS[tipo]||'#718096';
  const cColor=CARREIRA_COLORS[carreira]||'#718096';
  const fColor=FASE_COLORS[fase]||'#718096';
  document.getElementById('sp-badges').innerHTML=[
    buildBadge(hl(tipo,TIPO_LABELS),tColor),
    buildBadge(hl(carreira,CARREIRA_LABELS),cColor),
    buildBadge(fase,fColor),
    materia?buildBadge(hl(materia,MATERIA_LABELS),'#718096'):'',
    material&&material!=='nenhum'?buildBadge(hl(material,MATERIAL_LABELS),'#0694A2'):'',
  ].join('');

  const tSupLabel = r[15] ? (TIPOSUPORTE_LABELS[r[15]] || r[15]) : '—';
  document.getElementById('sp-meta').innerHTML=`
    <div class="sp-meta-item"><div class="sp-meta-label">Data</div><div class="sp-meta-value">${r[0]||'—'}</div></div>
    <div class="sp-meta-item"><div class="sp-meta-label">Status</div><div class="sp-meta-value">${r[8]||'—'}</div></div>
    <div class="sp-meta-item"><div class="sp-meta-label">Coordenador</div><div class="sp-meta-value">${r[9]||'—'}</div></div>
    <div class="sp-meta-item"><div class="sp-meta-label">Turma</div><div class="sp-meta-value" title="${r[7]||''}">${r[7]?(r[7].length>25?r[7].slice(0,25)+'…':r[7]):'—'}</div></div>
    <div class="sp-meta-item"><div class="sp-meta-label">Confiança IA</div><div class="sp-meta-value">${confBadge(r[10])}</div></div>
    <div class="sp-meta-item"><div class="sp-meta-label">Matéria</div><div class="sp-meta-value">${hl(r[5]||'',MATERIA_LABELS)||'—'}</div></div>
    <div class="sp-meta-item" style="grid-column:span 2"><div class="sp-meta-label">Tipo de Suporte (original)</div><div class="sp-meta-value" style="font-size:12px">${tSupLabel}</div></div>
  `;

  document.getElementById('sp-desc').textContent = r[11]||'Sem descrição';
  document.getElementById('sp-resp').textContent = r[12]||'Sem resposta';

  if(r[13]){
    document.getElementById('sp-nota-sec').style.display='block';
    document.getElementById('sp-nota').textContent=r[13];
  } else {
    document.getElementById('sp-nota-sec').style.display='none';
  }

  document.getElementById('side-panel').classList.add('open');
  document.getElementById('overlay').classList.add('show');

  // Highlight selected row
  const pageStart = explorerPage * PAGE_SIZE;
  document.querySelectorAll('#explorer-tbody tr').forEach((tr,i)=>{
    tr.classList.toggle('selected', pageStart+i===idx);
  });
}

function closePanel(){
  document.getElementById('side-panel').classList.remove('open');
  document.getElementById('overlay').classList.remove('show');
  selectedRow=-1;
}

// ============================================================
// 15. TAB SWITCHING
// ============================================================
function switchTab(idx){
  document.querySelectorAll('.content').forEach((el,i)=>el.classList.toggle('active',i===idx));
  document.querySelectorAll('.tab-btn').forEach((el,i)=>el.classList.toggle('active',i===idx));
  if(idx===1) renderNatureza();
  if(idx===2) renderCarreira();
  if(idx===3) renderExplorer();
}

// ============================================================
// 16. INIT
// ============================================================
function populateSelects(){
  // Top 30 turmas for filter
  const tc={};
  FD.forEach(r=>{ if(r[6]) tc[r[6]]=(tc[r[6]]||0)+1; });
  const top30 = Object.entries(tc).sort((a,b)=>b[1]-a[1]).slice(0,30);
  const turmaSelect = document.getElementById('f0-turma');
  top30.forEach(([t,v])=>{
    const o=document.createElement('option');
    o.value=t; o.textContent=`${t} (${v})`;
    turmaSelect.appendChild(o);
  });

  // Months
  const meses = [...new Set(FD.map(r=>r[8]).filter(x=>x))].sort();
  const mesSelect = document.getElementById('f0-mes');
  const mesLabels={'2026-01':'Jan 2026','2026-02':'Fev 2026','2026-03':'Mar 2026','2026-04':'Abr 2026','2026-05':'Mai 2026','2026-06':'Jun 2026'};
  meses.forEach(m=>{
    const o=document.createElement('option');
    o.value=m; o.textContent=mesLabels[m]||m;
    mesSelect.appendChild(o);
  });

  // Explorer tipo select
  const tipoSel = document.getElementById('f3-tipo');
  Object.entries(TIPO_LABELS).forEach(([k,v])=>{
    const o=document.createElement('option'); o.value=k; o.textContent=v; tipoSel.appendChild(o);
  });

  // Explorer materia select
  const matSel = document.getElementById('f3-materia');
  const allMat = [...new Set(FD.map(r=>r[4]).filter(x=>x))].sort();
  allMat.forEach(m=>{
    const o=document.createElement('option'); o.value=m; o.textContent=hl(m,MATERIA_LABELS); matSel.appendChild(o);
  });

  // Tipo de Suporte selects (Aba 0 e Aba 3)
  // collect unique values sorted by frequency
  const tsCount={};
  FD.forEach(r=>{ const k=r[11]; if(k) tsCount[k]=(tsCount[k]||0)+1; });
  const tsSorted = Object.entries(tsCount).sort((a,b)=>b[1]-a[1]);
  ['f0-tiposuporte','f3-tiposuporte'].forEach(id=>{
    const sel = document.getElementById(id);
    tsSorted.forEach(([k,v])=>{
      const o=document.createElement('option');
      o.value=k;
      o.textContent=(TIPOSUPORTE_LABELS[k]||k)+' ('+v+')';
      sel.appendChild(o);
    });
  });
}

// Color pills on load
document.querySelectorAll('.carreira-pills .pill[data-carreira]').forEach(p=>{
  const c = p.dataset.carreira;
  if(c && CARREIRA_COLORS[c]){
    p.addEventListener('mouseenter',()=>{ if(!p.classList.contains('active')){ p.style.borderColor=CARREIRA_COLORS[c]; p.style.color=CARREIRA_COLORS[c]; }});
    p.addEventListener('mouseleave',()=>{ if(!p.classList.contains('active')){ p.style.borderColor=''; p.style.color=''; }});
  }
});

populateSelects();
applyFilters();

</script>
</body>
</html>
"""

# Inject data
HTML = HTML.replace('__FILTER_DATA__', FILTER_JSON)
HTML = HTML.replace('__FULL_DATA__', FULL_JSON)

with open('dashboard_mege.html','w',encoding='utf-8') as f:
    f.write(HTML)

import os
size = os.path.getsize('dashboard_mege.html')
print(f'Done! File size: {size/1024:.1f} KB ({size/1024/1024:.2f} MB)')
